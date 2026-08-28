"""Rapor Üretici — Excel ve PDF dışa aktarma.

10 günlük dönem bazlı raporlar üretir:
- 1-10 / 11-20 / 21-ay sonu
- Her dönem için KDV dökümü
- Aylık özet
- Firma ismi desteği
- DejaVu Sans fontu (Türkçe karakter desteği)
"""
import os
import json
import logging
from datetime import date, datetime
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from utils.helpers import (
    get_10day_periods, format_turkish_lira, MONTH_NAMES_TR
)

logger = logging.getLogger(__name__)

# ========================================================================
# FONT KAYDI — Türkçe karakter desteği (ç, ş, ğ, ı, ö, ü, İ, Ş, Ç, Ğ)
# ========================================================================
FONT_DIR = "/usr/share/fonts/truetype/dejavu"
_tried_register = False

def _register_fonts():
    global _tried_register
    if _tried_register:
        return
    _tried_register = True
    try:
        pdfmetrics.registerFont(TTFont("DejaVuSans", f"{FONT_DIR}/DejaVuSans.ttf"))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", f"{FONT_DIR}/DejaVuSans-Bold.ttf"))
        logger.info("DejaVu Sans fontları kaydedildi (Türkçe karakter desteği aktif)")
    except Exception as e:
        logger.warning(f"DejaVu font yüklenemedi, varsayılan font kullanılacak: {e}")

_register_fonts()

# PDF font adları
PDF_FONT = "DejaVuSans"
PDF_FONT_BOLD = "DejaVuSans-Bold"

# ============================================================================
# EXCEL STIL TANIMLARI
# ============================================================================
EXCEL_FONT = "DejaVu Sans"
EXCEL_FONT_BOLD = "DejaVu Sans"

HEADER_FILL = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
HEADER_FONT = Font(name=EXCEL_FONT_BOLD, bold=True, color="FFFFFF", size=10)
TOTAL_FILL = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
TOTAL_FONT = Font(name=EXCEL_FONT_BOLD, bold=True, size=10)
NORMAL_FONT = Font(name=EXCEL_FONT, size=10)
TITLE_FONT = Font(name=EXCEL_FONT_BOLD, bold=True, size=14)
SUBTITLE_FONT = Font(name=EXCEL_FONT_BOLD, bold=True, size=11, color="64748B")

# Görünür kenarlık rengi (koyu gri)
THIN_BORDER = Border(
    left=Side(style="thin", color="94A3B8"),
    right=Side(style="thin", color="94A3B8"),
    top=Side(style="thin", color="94A3B8"),
    bottom=Side(style="thin", color="94A3B8"),
)

# KDV oranı etiketleri
VAT_RATE_LABELS = {0: "%0", 1: "%1", 10: "%10", 20: "%20"}

# Ortak KDV oranları — her zaman sütun olarak göster
COMMON_VAT_RATES = [0, 1, 10, 20]


def _get_vat_rate_label(exp: dict) -> str:
    """Giderin KDV oranını gösteren etiket üret.
    
    Çoklu KDV oranı varsa 'Muhtelif Oranlar' döndür.
    Tek oran varsa '%XX' formatında döndür.
    """
    vat_items = exp.get("vat_items")
    if vat_items and len(vat_items) > 1:
        return "Muhtelif"
    
    if vat_items and len(vat_items) == 1:
        rate = vat_items[0].get("vat_rate", 0)
        return VAT_RATE_LABELS.get(int(rate), f"%{int(rate)}")
    
    vat_rate = exp.get("vat_rate")
    if vat_rate is not None:
        return VAT_RATE_LABELS.get(int(vat_rate), f"%{int(vat_rate)}")
    
    return "-"


def _get_vat_detail_text(exp: dict) -> str:
    """Giderin KDV detay metnini üret.
    
    Çoklu KDV varsa her birini listeler:
    "%10: 155,45₺ | %20: 0,17₺"
    """
    vat_items = exp.get("vat_items")
    if vat_items and len(vat_items) > 1:
        parts = []
        for item in vat_items:
            rate = int(item.get("vat_rate", 0))
            amount = item.get("vat_amount", 0) or 0
            parts.append(f"%{rate}: {format_turkish_lira(amount)}")
        return " | ".join(parts)
    
    vat_amount = exp.get("vat_amount") or 0
    return format_turkish_lira(vat_amount)


def _get_vat_items(exp: dict) -> List[dict]:
    """Giderin KDV satırlarını döndür (tek veya çoklu)."""
    vat_items = exp.get("vat_items")
    if vat_items and len(vat_items) > 0:
        return vat_items
    
    vat_rate = exp.get("vat_rate") or 0
    return [{
        "vat_rate": vat_rate,
        "net_amount": exp.get("net_amount") or 0,
        "vat_amount": exp.get("vat_amount") or 0,
        "total_amount": exp.get("total_amount") or 0,
    }]


def _format_vat_cell(exp: dict) -> str:
    """KDV hücresi için formatlı metin üret."""
    vat_items = exp.get("vat_items")
    if vat_items and len(vat_items) > 1:
        return "Muhtelif Oranlar"
    
    vat_rate = exp.get("vat_rate")
    if vat_rate is not None:
        return f"%{int(vat_rate)}"
    return "-"


# ============================================================================
# EXCEL RAPOR — Referans formata uygun
# ============================================================================

def generate_excel_report(
    periods_data: List[Dict],
    year: int,
    month: int,
    output_path: str,
    company_name: str = "",
) -> str:
    """Aylık 10 günlük dönem Excel raporu — A4 yatay format.

    Format:
    - Satır 1: (Firma İsmi) X.Ay (AyAdı) Yıl DÖNEMİ
    - Satır 4: Periyodik Toplam | 1-10 | 11-20 | 21-aysonu | Genel Toplam
    - Satır 5: KDV oranları (alt sütunlar, %0 hariç)
    - Satır 6+: Gider tutarları
    - Alt satırlar: KDV Dahil Tutar, Toplam Matrah, Toplam KDV
    """
    wb = Workbook()
    month_name = MONTH_NAMES_TR[month]

    # ====================================================================
    # SAYFA 1: ÖZET
    # ====================================================================
    ws = wb.active
    ws.title = "Özet"

    # Firma ismi + dönem başlığı
    company_part = f"{company_name} " if company_name else ""
    title_text = f"{company_part}{month}.Ay ({month_name}) {year} DÖNEMİ"
    ws.merge_cells("A1:K1")
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Aktif KDV oranlarını bul — tüm dönemlerdeki giderlerden (%0 hariç)
    all_vat_rates = set()
    for period_data in periods_data:
        for exp in period_data.get("expenses", []):
            items = _get_vat_items(exp)
            for item in items:
                rate = int(item.get("vat_rate", 0))
                all_vat_rates.add(rate)

    # Aktif KDV oranlarını bul — verideki oranları göster (0 hariç)
    active_rates = sorted(set(r for r in all_vat_rates if r > 0))
    if not active_rates:
        active_rates = [10, 20]  # Varsayılan olarak %10 ve %20 göster

    # ====================================================================
    # SÜTUN YAPISI: A | [dönemlerdeki KDV oranları] | Genel Toplam
    # Dönem toplam sütunu YOK — sadece genel toplam var
    # ====================================================================
    
    # Her dönem için sütun haritası: {period_idx: {rate: col}}
    period_cols = {}  # {period_idx: {rate: col}}
    col = 2  # A sütunu"A", B'den başla

    for p_idx, period_data in enumerate(periods_data):
        period_cols[p_idx] = {}
        for rate in active_rates:
            period_cols[p_idx][rate] = col
            col += 1

    # Genel toplam sütunu
    grand_total_col = col

    # ====================================================================
    # SATIR 4: Dönem başlıkları
    # ====================================================================
    row = 4
    ws.cell(row=row, column=1, value="Periyodik Toplam").font = TOTAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER

    for p_idx, period_data in enumerate(periods_data):
        period = period_data["period"]
        # Dönem etiketi: "1-10", "11-20", "21-31" gibi
        period_label = period["label"].split(" ")[0]  # "1-10 Ağustos 2026" → "1-10"

        start_col = period_cols[p_idx][active_rates[0]]
        end_col = period_cols[p_idx][active_rates[-1]]

        if len(active_rates) > 1:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        ws.cell(row=row, column=start_col, value=period_label).font = TOTAL_FONT
        ws.cell(row=row, column=start_col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=start_col).border = THIN_BORDER

        for rate in active_rates:
            c = period_cols[p_idx][rate]
            ws.cell(row=row, column=c).border = THIN_BORDER

    # Genel toplam başlığı
    ws.cell(row=row, column=grand_total_col, value="Genel\nToplam").font = TOTAL_FONT
    ws.cell(row=row, column=grand_total_col).alignment = Alignment(horizontal="center", wrap_text=True)
    ws.cell(row=row, column=grand_total_col).border = THIN_BORDER

    # ====================================================================
    # SATIR 5: KDV oranları
    # ====================================================================
    row = 5
    ws.cell(row=row, column=1).border = THIN_BORDER
    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            ws.cell(row=row, column=c, value=f"%{rate}").font = NORMAL_FONT
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=c).border = THIN_BORDER
    ws.cell(row=row, column=grand_total_col).border = THIN_BORDER

    # ====================================================================
    # SATIR 6+: Gider verileri
    # ====================================================================
    data_start_row = 6
    row = 6
    max_expenses = max(len(p.get("expenses", [])) for p in periods_data) if periods_data else 0

    for exp_idx in range(max_expenses):
        ws.cell(row=row, column=1).border = THIN_BORDER

        for p_idx, period_data in enumerate(periods_data):
            expenses = period_data.get("expenses", [])
            if exp_idx < len(expenses):
                exp = expenses[exp_idx]
                items = _get_vat_items(exp)

                for rate in active_rates:
                    c = period_cols[p_idx][rate]
                    rate_items = [i for i in items if int(i.get("vat_rate", 0)) == rate]
                    if rate_items:
                        amount = sum(i.get("total_amount", 0) or 0 for i in rate_items)
                    else:
                        amount = 0

                    cell = ws.cell(row=row, column=c, value=amount if amount > 0 else None)
                    cell.font = NORMAL_FONT
                    cell.number_format = '#,##0.00'
                    cell.border = THIN_BORDER
            else:
                for rate in active_rates:
                    c = period_cols[p_idx][rate]
                    ws.cell(row=row, column=c).border = THIN_BORDER

        # Genel toplam sütunu — tüm sütunların toplamı (formül)
        all_col_refs = []
        for p_idx in range(len(periods_data)):
            for rate in active_rates:
                c = period_cols[p_idx][rate]
                all_col_refs.append(f"{get_column_letter(c)}{row}")
        formula = "=" + "+".join(all_col_refs)
        cell = ws.cell(row=row, column=grand_total_col, value=formula)
        cell.font = NORMAL_FONT
        cell.number_format = '#,##0.00'
        cell.border = THIN_BORDER

        row += 1

    data_end_row = row - 1

    # Boş satır
    row += 1

    # ====================================================================
    # KDV DAHİL TUTAR satırı
    # ====================================================================
    ws.cell(row=row, column=1, value="KDV Dahil Tutar").font = TOTAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).fill = TOTAL_FILL

    kdv_dahil_row = row

    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            col_letter = get_column_letter(c)
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.font = TOTAL_FONT
            cell.number_format = '#,##0.00'
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER

    # Genel toplam — tüm sütun KDV dahil toplamlarının toplamı
    grand_refs = []
    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            grand_refs.append(f"{get_column_letter(c)}{row}")
    cell = ws.cell(row=row, column=grand_total_col, value="=" + "+".join(grand_refs))
    cell.font = TOTAL_FONT
    cell.number_format = '#,##0.00'
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    row += 1

    # ====================================================================
    # TOPLAM MATRAH satırı
    # ====================================================================
    ws.cell(row=row, column=1, value="Toplam Matrah").font = TOTAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).fill = TOTAL_FILL

    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            col_letter = get_column_letter(c)
            if rate > 0:
                multiplier = 1 + rate / 100
                formula = f"={col_letter}{kdv_dahil_row}/{multiplier}"
            else:
                formula = f"={col_letter}{kdv_dahil_row}"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.font = TOTAL_FONT
            cell.number_format = '#,##0.00'
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER

    # Genel toplam matrah
    grand_refs = []
    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            grand_refs.append(f"{get_column_letter(c)}{row}")
    cell = ws.cell(row=row, column=grand_total_col, value="=" + "+".join(grand_refs))
    cell.font = TOTAL_FONT
    cell.number_format = '#,##0.00'
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    row += 1

    # ====================================================================
    # TOPLAM KDV satırı
    # ====================================================================
    ws.cell(row=row, column=1, value="Toplam KDV").font = TOTAL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).fill = TOTAL_FILL

    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            col_letter = get_column_letter(c)
            formula = f"={col_letter}{kdv_dahil_row}-{col_letter}{row - 1}"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.font = TOTAL_FONT
            cell.number_format = '#,##0.00'
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER

    # Genel toplam KDV
    grand_refs = []
    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            grand_refs.append(f"{get_column_letter(c)}{row}")
    cell = ws.cell(row=row, column=grand_total_col, value="=" + "+".join(grand_refs))
    cell.font = TOTAL_FONT
    cell.number_format = '#,##0.00'
    cell.fill = TOTAL_FILL
    cell.border = THIN_BORDER

    # ====================================================================
    # SÜTUN GENİŞLİKLERİ VE A4 YATAY BASKI AYARLARI
    # ====================================================================
    ws.column_dimensions["A"].width = 18
    for p_idx in range(len(periods_data)):
        for rate in active_rates:
            c = period_cols[p_idx][rate]
            ws.column_dimensions[get_column_letter(c)].width = 13
    ws.column_dimensions[get_column_letter(grand_total_col)].width = 15

    # A4 yatay baskı ayarları
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Yükseklik otomatik (gider sayısına göre)
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    # Yazdırma alanı
    last_col_letter = get_column_letter(grand_total_col)
    ws.print_area = f"A1:{last_col_letter}{row}"

    # ====================================================================
    # HER DÖNEM İÇİN AYRI SAYFA (detay)
    # ====================================================================
    detail_headers = ["Tarih", "İş Yeri", "Açıklama", "KDV%", "Matrah (₺)", "KDV (₺)", "Toplam (₺)"]

    for p_idx, period_data in enumerate(periods_data):
        period = period_data["period"]
        expenses = period_data["expenses"]

        ws_detail = wb.create_sheet(title=period["label"][:31])

        # Başlık
        ws_detail.merge_cells("A1:G1")
        ws_detail["A1"] = f"{company_part}Dönem: {period['label']}"
        ws_detail["A1"].font = TITLE_FONT
        ws_detail["A1"].alignment = Alignment(horizontal="center")

        # Sütun başlıkları
        for col_idx, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=3, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Veriler
        detail_row = 4
        period_total = 0
        period_net = 0
        period_vat = 0

        for exp in expenses:
            ws_detail.cell(row=detail_row, column=1, value=exp.get("receipt_date", "")).font = NORMAL_FONT
            ws_detail.cell(row=detail_row, column=2, value=exp.get("vendor_name", "")).font = NORMAL_FONT
            ws_detail.cell(row=detail_row, column=3, value=exp.get("description", "")).font = NORMAL_FONT

            # KDV hücresi — çoklu KDV dökümü
            vat_text = _format_vat_cell(exp)
            ws_detail.cell(row=detail_row, column=4, value=vat_text).font = NORMAL_FONT

            ws_detail.cell(row=detail_row, column=5, value=exp.get("net_amount") or 0).font = NORMAL_FONT
            ws_detail.cell(row=detail_row, column=6, value=exp.get("vat_amount") or 0).font = NORMAL_FONT
            ws_detail.cell(row=detail_row, column=7, value=exp.get("total_amount", 0)).font = NORMAL_FONT

            for c in range(1, 8):
                ws_detail.cell(row=detail_row, column=c).border = THIN_BORDER
                if c >= 5:
                    ws_detail.cell(row=detail_row, column=c).number_format = '#,##0.00'

            period_total += exp.get("total_amount", 0)
            period_net += exp.get("net_amount") or 0
            period_vat += exp.get("vat_amount") or 0
            detail_row += 1

            # Çoklu KDV varsa ek satır detay
            vat_items = exp.get("vat_items")
            if vat_items and len(vat_items) > 1:
                for item in vat_items:
                    rate = int(item.get("vat_rate", 0))
                    ws_detail.cell(row=detail_row, column=4, value=f"  ↳ %{rate}").font = Font(name=EXCEL_FONT, size=9, color="64748B")
                    ws_detail.cell(row=detail_row, column=5, value=item.get("net_amount") or 0).font = Font(name=EXCEL_FONT, size=9, color="64748B")
                    ws_detail.cell(row=detail_row, column=6, value=item.get("vat_amount") or 0).font = Font(name=EXCEL_FONT, size=9, color="64748B")
                    ws_detail.cell(row=detail_row, column=7, value=item.get("total_amount") or 0).font = Font(name=EXCEL_FONT, size=9, color="64748B")
                    for c in range(1, 8):
                        ws_detail.cell(row=detail_row, column=c).border = THIN_BORDER
                        if c >= 5:
                            ws_detail.cell(row=detail_row, column=c).number_format = '#,##0.00'
                    detail_row += 1

        # Dönem toplamı
        detail_row += 1
        ws_detail.cell(row=detail_row, column=1, value="DÖNEM TOPLAMI").font = TOTAL_FONT
        ws_detail.cell(row=detail_row, column=5, value=period_net).font = TOTAL_FONT
        ws_detail.cell(row=detail_row, column=6, value=period_vat).font = TOTAL_FONT
        ws_detail.cell(row=detail_row, column=7, value=period_total).font = TOTAL_FONT

        for c in range(1, 8):
            cell = ws_detail.cell(row=detail_row, column=c)
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
            if c >= 5:
                cell.number_format = '#,##0.00'

        # Sütun genişlikleri
        ws_detail.column_dimensions["A"].width = 14
        ws_detail.column_dimensions["B"].width = 25
        ws_detail.column_dimensions["C"].width = 30
        ws_detail.column_dimensions["D"].width = 16
        ws_detail.column_dimensions["E"].width = 16
        ws_detail.column_dimensions["F"].width = 14
        ws_detail.column_dimensions["G"].width = 16

    # Kaydet
    wb.save(output_path)
    logger.info(f"Excel raporu kaydedildi: {output_path}")
    return output_path


# ============================================================================
# PDF RAPOR — DejaVu Sans fontu + KDV dökümü
# ============================================================================

def generate_pdf_report(
    periods_data: List[Dict],
    year: int,
    month: int,
    output_path: str,
    company_name: str = "",
) -> str:
    """Aylık 10 günlük dönem PDF raporu üretir.
    
    DejaVu Sans fontu ile Türkçe karakter desteği.
    Çoklu KDV oranları detaylı gösterilir.
    """
    month_name = MONTH_NAMES_TR[month]

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    # Özel stiller — DejaVu Sans ile Türkçe karakter desteği
    title_style = ParagraphStyle(
        "ReportTitle",
        fontName=PDF_FONT_BOLD,
        fontSize=18,
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        fontName=PDF_FONT,
        fontSize=10,
        textColor=colors.HexColor("#64748B"),
        alignment=TA_CENTER,
        spaceAfter=12,
    )
    period_title_style = ParagraphStyle(
        "PeriodTitle",
        fontName=PDF_FONT_BOLD,
        fontSize=13,
        textColor=colors.HexColor("#0284C7"),
        spaceAfter=8,
        spaceBefore=16,
    )
    normal_style = ParagraphStyle(
        "NormalTR",
        fontName=PDF_FONT,
        fontSize=9,
    )
    bold_style = ParagraphStyle(
        "BoldTR",
        fontName=PDF_FONT_BOLD,
        fontSize=9,
    )
    right_style = ParagraphStyle(
        "RightTR",
        fontName=PDF_FONT,
        fontSize=9,
        alignment=TA_RIGHT,
    )
    small_style = ParagraphStyle(
        "SmallTR",
        fontName=PDF_FONT,
        fontSize=7,
        textColor=colors.HexColor("#64748B"),
    )
    summary_title_style = ParagraphStyle(
        "SummaryTitle",
        fontName=PDF_FONT_BOLD,
        fontSize=14,
        textColor=colors.HexColor("#0284C7"),
        spaceBefore=20,
        spaceAfter=12,
    )

    elements = []

    # ========================================================================
    # BAŞLIK
    # ========================================================================
    company_part = f"{company_name} — " if company_name else ""
    elements.append(Paragraph(f"{company_part}Aylık Gider Raporu", title_style))
    elements.append(Paragraph(f"{month_name} {year}", subtitle_style))
    elements.append(Paragraph(
        f"Oluşturma: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        subtitle_style,
    ))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor("#E2E8F0"),
        spaceAfter=12,
    ))

    # ========================================================================
    # HER DÖNEM İÇİN TABLO
    # ========================================================================
    for period_data in periods_data:
        period = period_data["period"]
        expenses = period_data["expenses"]
        totals = period_data["totals"]

        elements.append(Paragraph(f"Dönem: {period['label']}", period_title_style))

        # Tablo başlığı
        table_data = [[
            Paragraph("Tarih", bold_style),
            Paragraph("İş Yeri", bold_style),
            Paragraph("Açıklama", bold_style),
            Paragraph("KDV%", bold_style),
            Paragraph("Matrah", bold_style),
            Paragraph("KDV", bold_style),
            Paragraph("Toplam", bold_style),
        ]]

        for exp in expenses:
            # KDV hücresi — çoklu ise detay göster
            vat_text = _format_vat_cell(exp)
            vat_detail = _get_vat_detail_text(exp)

            if vat_detail and vat_text == "Muhtelif Oranlar":
                vat_cell = Paragraph(
                    f"<b>{vat_text}</b><br/><font size=7 color='#64748B'>{vat_detail}</font>",
                    normal_style,
                )
            else:
                vat_cell = Paragraph(vat_text, normal_style)

            table_data.append([
                Paragraph(str(exp.get("receipt_date", "")), normal_style),
                Paragraph(str(exp.get("vendor_name", "")), normal_style),
                Paragraph(str(exp.get("description", "")[:40]), normal_style),
                vat_cell,
                Paragraph(format_turkish_lira(exp.get("net_amount") or 0), normal_style),
                Paragraph(format_turkish_lira(exp.get("vat_amount") or 0), normal_style),
                Paragraph(format_turkish_lira(exp.get("total_amount", 0)), normal_style),
            ])

        # Toplam satırı
        table_data.append([
            "",
            Paragraph("<b>TOPLAM</b>", bold_style),
            "",
            "",
            Paragraph(format_turkish_lira(totals["net"]), bold_style),
            Paragraph(format_turkish_lira(totals["vat"]), bold_style),
            Paragraph(format_turkish_lira(totals["total"]), bold_style),
        ])

        # Tabloyu oluştur
        col_widths = [2.2*cm, 4.5*cm, 5*cm, 3.5*cm, 3*cm, 2.5*cm, 3*cm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Stil
        table_style = TableStyle([
            # Başlık satırı
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284C7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),

            # Veri satırları
            ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (4, 1), (6, -1), "RIGHT"),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),

            # Toplam satırı
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0F9FF")),
            ("FONTNAME", (0, -1), (-1, -1), PDF_FONT_BOLD),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#0284C7")),

            # Kenarlıklar
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#94A3B8")),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#64748B")),
        ])
        table.setStyle(table_style)
        elements.append(table)
        elements.append(Spacer(1, 0.5 * cm))

    # ========================================================================
    # AYLIK ÖZET
    # ========================================================================
    elements.append(HRFlowable(
        width="100%", thickness=1, color=colors.HexColor("#E2E8F0"),
        spaceAfter=8, spaceBefore=16,
    ))
    elements.append(Paragraph("Aylık Toplam Özet", summary_title_style))

    grand_total = sum(p["totals"]["total"] for p in periods_data)
    grand_net = sum(p["totals"]["net"] for p in periods_data)
    grand_vat = sum(p["totals"]["vat"] for p in periods_data)
    grand_count = sum(p["totals"]["count"] for p in periods_data)

    summary_data = [
        [
            Paragraph("Toplam Kayıt", ParagraphStyle("sh", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Toplam Matrah", ParagraphStyle("sh2", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Toplam KDV", ParagraphStyle("sh3", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Genel Toplam", ParagraphStyle("sh4", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
        ],
        [
            Paragraph(str(grand_count), ParagraphStyle("sv", fontName=PDF_FONT_BOLD, fontSize=11, alignment=TA_CENTER)),
            Paragraph(format_turkish_lira(grand_net), ParagraphStyle("sv2", fontName=PDF_FONT_BOLD, fontSize=11, alignment=TA_CENTER)),
            Paragraph(format_turkish_lira(grand_vat), ParagraphStyle("sv3", fontName=PDF_FONT_BOLD, fontSize=11, alignment=TA_CENTER)),
            Paragraph(format_turkish_lira(grand_total), ParagraphStyle("sv4", fontName=PDF_FONT_BOLD, fontSize=11, alignment=TA_CENTER)),
        ],
    ]

    summary_table = Table(summary_data, colWidths=[4*cm, 4.5*cm, 3.5*cm, 4*cm])
    summary_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284C7")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F0F9FF")),
        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#0284C7")),
        ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#0284C7")),
    ])
    summary_table.setStyle(summary_style)
    elements.append(summary_table)

    # ========================================================================
    # KDV DETAY DÖKÜMÜ (PDF)
    # ========================================================================
    # Tüm dönemlerdeki KDV oranlarını topla — herbir fişi doğru orana分离 et
    vat_rate_totals = {}
    for period_data in periods_data:
        for exp in period_data.get("expenses", []):
            items = _get_vat_items(exp)
            for item in items:
                rate = int(item.get("vat_rate", 0))
                if rate not in vat_rate_totals:
                    vat_rate_totals[rate] = {"net": 0, "vat": 0, "total": 0, "count": 0}
                vat_rate_totals[rate]["net"] += item.get("net_amount", 0) or 0
                vat_rate_totals[rate]["vat"] += item.get("vat_amount", 0) or 0
                vat_rate_totals[rate]["total"] += item.get("total_amount", 0) or 0
                vat_rate_totals[rate]["count"] += 1

    if len(vat_rate_totals) >= 1:  # En az bir KDV oranı varsa detay göster
        elements.append(Spacer(1, 0.8 * cm))
        elements.append(Paragraph("KDV Oranı Dökümü", summary_title_style))

        vat_detail_data = [[
            Paragraph("KDV Oranı", ParagraphStyle("vh", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Kayıt", ParagraphStyle("vh2", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Matrah", ParagraphStyle("vh3", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("KDV", ParagraphStyle("vh4", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
            Paragraph("Toplam", ParagraphStyle("vh5", fontName=PDF_FONT_BOLD, fontSize=10, textColor=colors.white, alignment=TA_CENTER)),
        ]]

        for rate in sorted(vat_rate_totals.keys()):
            g = vat_rate_totals[rate]
            vat_detail_data.append([
                Paragraph(f"%{rate}", ParagraphStyle("vr", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
                Paragraph(str(g["count"]), ParagraphStyle("vc", fontName=PDF_FONT, fontSize=10, alignment=TA_CENTER)),
                Paragraph(format_turkish_lira(g["net"]), ParagraphStyle("vn", fontName=PDF_FONT, fontSize=10, alignment=TA_CENTER)),
                Paragraph(format_turkish_lira(g["vat"]), ParagraphStyle("vv", fontName=PDF_FONT, fontSize=10, alignment=TA_CENTER)),
                Paragraph(format_turkish_lira(g["total"]), ParagraphStyle("vt", fontName=PDF_FONT, fontSize=10, alignment=TA_CENTER)),
            ])

        # Toplam satırı
        vat_detail_data.append([
            Paragraph("<b>TOPLAM</b>", ParagraphStyle("vtl", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"<b>{grand_count}</b>", ParagraphStyle("vtc", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"<b>{format_turkish_lira(grand_net)}</b>", ParagraphStyle("vtn", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"<b>{format_turkish_lira(grand_vat)}</b>", ParagraphStyle("vtv", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"<b>{format_turkish_lira(grand_total)}</b>", ParagraphStyle("vtt", fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER)),
        ])

        vat_table = Table(vat_detail_data, colWidths=[3*cm, 2.5*cm, 4*cm, 3.5*cm, 4*cm])
        vat_table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#92400E")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF3C7")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#92400E")),
            ("BOX", (0, 0), (-1, -1), 1.5, colors.HexColor("#92400E")),
        ])
        vat_table.setStyle(vat_table_style)
        elements.append(vat_table)

    # PDF'i oluştur
    doc.build(elements)
    logger.info(f"PDF raporu kaydedildi: {output_path}")
    return output_path
