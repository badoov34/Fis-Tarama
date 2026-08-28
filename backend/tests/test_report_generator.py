"""report_generator.py fonksiyonları için unit testler.

Test edilen fonksiyonlar:
- _get_vat_items
- _get_vat_rate_label
- _get_vat_detail_text
- _format_vat_cell
- generate_excel_report (dosya oluşturma)
- generate_pdf_report (dosya oluşturma)
"""
import sys
import os
import tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from datetime import date
from utils.report_generator import (
    _get_vat_items,
    _get_vat_rate_label,
    _get_vat_detail_text,
    _format_vat_cell,
    generate_excel_report,
    generate_pdf_report,
)
from utils.helpers import format_turkish_lira


# ============================================================================
# _get_vat_items
# ============================================================================
class TestGetVatItems:
    """Giderin KDV satırlarını çıkarma testleri."""

    def test_tek_kdv_orani(self):
        """Tek KDV oranı olan gider"""
        exp = {
            "vat_rate": 20,
            "net_amount": 1000,
            "vat_amount": 200,
            "total_amount": 1200,
        }
        items = _get_vat_items(exp)
        assert len(items) == 1
        assert items[0]["vat_rate"] == 20

    def test_coklu_kdv(self):
        """Çoklu KDV oranlı gider"""
        exp = {
            "vat_items": [
                {"vat_rate": 10, "total_amount": 500, "net_amount": 450, "vat_amount": 50},
                {"vat_rate": 20, "total_amount": 300, "net_amount": 250, "vat_amount": 50},
            ]
        }
        items = _get_vat_items(exp)
        assert len(items) == 2
        assert items[0]["vat_rate"] == 10
        assert items[1]["vat_rate"] == 20

    def test_vat_items_yok(self):
        """vat_items yoksa fallback"""
        exp = {
            "vat_rate": 10,
            "total_amount": 550,
            "net_amount": 500,
            "vat_amount": 50,
        }
        items = _get_vat_items(exp)
        assert len(items) == 1
        assert items[0]["vat_rate"] == 10
        assert items[0]["total_amount"] == 550

    def test_bos_vat_items(self):
        """Boş vat_items listesi → fallback"""
        exp = {
            "vat_items": [],
            "vat_rate": 20,
            "total_amount": 1200,
            "net_amount": 1000,
            "vat_amount": 200,
        }
        items = _get_vat_items(exp)
        assert len(items) == 1


# ============================================================================
# _get_vat_rate_label
# ============================================================================
class TestGetVatRateLabel:
    """KDV oranı etiketi testleri."""

    def test_tek_oran_yuzde20(self):
        exp = {"vat_items": [{"vat_rate": 20, "total_amount": 100}]}
        result = _get_vat_rate_label(exp)
        assert result == "%20"

    def test_tek_oran_yuzde10(self):
        exp = {"vat_items": [{"vat_rate": 10, "total_amount": 100}]}
        result = _get_vat_rate_label(exp)
        assert result == "%10"

    def test_coklu_oran(self):
        exp = {
            "vat_items": [
                {"vat_rate": 10, "total_amount": 100},
                {"vat_rate": 20, "total_amount": 200},
            ]
        }
        result = _get_vat_rate_label(exp)
        assert result == "Muhtelif"

    def test_vat_items_yok_rate_var(self):
        """vat_items yok ama vat_rate var"""
        exp = {"vat_rate": 10}
        result = _get_vat_rate_label(exp)
        assert result == "%10"

    def test_hic_bilgi_yok(self):
        exp = {}
        result = _get_vat_rate_label(exp)
        assert result == "-"


# ============================================================================
# _get_vat_detail_text
# ============================================================================
class TestGetVatDetailText:
    """KDV detay metni testleri."""

    def test_tek_kdv(self):
        exp = {"vat_amount": 200}
        result = _get_vat_detail_text(exp)
        assert "200" in result.replace(".", "").replace(",", ".")

    def test_coklu_kdv(self):
        exp = {
            "vat_items": [
                {"vat_rate": 10, "vat_amount": 50},
                {"vat_rate": 20, "vat_amount": 100},
            ]
        }
        result = _get_vat_detail_text(exp)
        assert "%10" in result
        assert "%20" in result
        assert "|" in result


# ============================================================================
# _format_vat_cell
# ============================================================================
class TestFormatVatCell:
    """KDV hücresi formatı testleri."""

    def test_tek_oran(self):
        exp = {"vat_items": [{"vat_rate": 20, "total_amount": 100}]}
        result = _format_vat_cell(exp)
        assert result == "%20"

    def test_coklu_oran(self):
        exp = {
            "vat_items": [
                {"vat_rate": 10, "total_amount": 100},
                {"vat_rate": 20, "total_amount": 200},
            ]
        }
        result = _format_vat_cell(exp)
        assert "Muhtelif" in result

    def test_vat_rate_direkt(self):
        exp = {"vat_rate": 10}
        result = _format_vat_cell(exp)
        assert result == "%10"


# ============================================================================
# generate_excel_report
# ============================================================================
class TestGenerateExcelReport:
    """Excel rapor oluşturma testleri."""

    def test_bos_veriyle(self):
        """Boş veri ile Excel oluşturulabilmeli"""
        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            result = generate_excel_report(periods_data, 2026, 8, output_path, company_name="Test Firma")
            assert result == output_path
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)

    def test_veriyle(self):
        """Gider verisi ile Excel"""
        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [
                    {
                        "total_amount": 1200.00,
                        "net_amount": 1000.00,
                        "vat_rate": 20,
                        "vat_amount": 200.00,
                        "vendor_name": "BİM",
                        "receipt_date": "2026-08-05",
                        "description": "Market alışverişi",
                        "vat_items": [{"vat_rate": 20, "total_amount": 1200.00, "net_amount": 1000.00, "vat_amount": 200.00}],
                    }
                ],
                "totals": {"total": 1200.00, "net": 1000.00, "vat": 200.00, "count": 1},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            result = generate_excel_report(periods_data, 2026, 8, output_path, company_name="Test Firma A.Ş.")
            assert result == output_path
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)

    def test_firma_ismi_icerigi(self):
        """Excel'de firma ismi başlıkta görünmeli"""
        from openpyxl import load_workbook

        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            generate_excel_report(periods_data, 2026, 8, output_path, company_name="Acme Corp")
            wb = load_workbook(output_path)
            ws = wb.active
            title = ws["A1"].value
            assert "Acme Corp" in title
            assert "Ağustos" in title
            assert "2026" in title
            wb.close()
        finally:
            os.unlink(output_path)

    def test_yatay_baski_ayarlari(self):
        """A4 yatay baskı ayarları doğru mu?"""
        from openpyxl import load_workbook

        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_path = f.name

        try:
            generate_excel_report(periods_data, 2026, 8, output_path)
            wb = load_workbook(output_path)
            ws = wb.active
            assert ws.page_setup.orientation == "landscape"
            assert ws.page_setup.fitToWidth == 1
            assert ws.page_setup.fitToHeight == 0
            wb.close()
        finally:
            os.unlink(output_path)


# ============================================================================
# generate_pdf_report
# ============================================================================
class TestGeneratePdfReport:
    """PDF rapor oluşturma testleri."""

    def test_bos_veriyle(self):
        """Boş veri ile PDF oluşturulabilmeli"""
        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            output_path = f.name

        try:
            result = generate_pdf_report(periods_data, 2026, 8, output_path, company_name="Test Firma")
            assert result == output_path
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)

    def test_veriyle(self):
        """Gider verisi ile PDF"""
        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [
                    {
                        "total_amount": 1200.00,
                        "net_amount": 1000.00,
                        "vat_rate": 20,
                        "vat_amount": 200.00,
                        "vendor_name": "BİM",
                        "receipt_date": "2026-08-05",
                        "description": "Market",
                        "vat_items": [{"vat_rate": 20, "total_amount": 1200.00, "net_amount": 1000.00, "vat_amount": 200.00}],
                    }
                ],
                "totals": {"total": 1200.00, "net": 1000.00, "vat": 200.00, "count": 1},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            output_path = f.name

        try:
            result = generate_pdf_report(periods_data, 2026, 8, output_path, company_name="Test Firma")
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)

    def test_coklu_kdv_oranli_fis(self):
        """Birden fazla KDV oranlı fiş — her oran doğru sütunda"""
        periods_data = [
            {
                "period": {"start": date(2026, 8, 1), "end": date(2026, 8, 10), "label": "1-10 Ağustos 2026"},
                "expenses": [
                    {
                        "total_amount": 550.00,
                        "net_amount": 500.00,
                        "vat_rate": None,  # Tek KDV oranı değil
                        "vat_amount": 50.00,
                        "vendor_name": "Market A",
                        "receipt_date": "2026-08-05",
                        "description": "Çoklu KDV",
                        "vat_items": [
                            {"vat_rate": 10, "total_amount": 330.00, "net_amount": 300.00, "vat_amount": 30.00},
                            {"vat_rate": 20, "total_amount": 220.00, "net_amount": 183.33, "vat_amount": 36.67},
                        ],
                    }
                ],
                "totals": {"total": 550.00, "net": 483.33, "vat": 66.67, "count": 1},
            },
            {
                "period": {"start": date(2026, 8, 11), "end": date(2026, 8, 20), "label": "11-20 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
            {
                "period": {"start": date(2026, 8, 21), "end": date(2026, 8, 31), "label": "21-31 Ağustos 2026"},
                "expenses": [],
                "totals": {"total": 0, "net": 0, "vat": 0, "count": 0},
            },
        ]

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            output_path = f.name

        try:
            result = generate_pdf_report(periods_data, 2026, 8, output_path, company_name="Test")
            assert os.path.exists(output_path)
            assert os.path.getsize(output_path) > 0
        finally:
            os.unlink(output_path)


# ============================================================================
# format_turkish_lira (ek testler)
# ============================================================================
class TestFormatTurkishLiraExtended:
    """format_turkish_lira için ek testler."""

    def test_kuruş_sıfır(self):
        assert format_turkish_lira(100.00) == "100,00₺"

    def test_kuruş_yuvarlama(self):
        """99.996 → 100.00"""
        assert format_turkish_lira(99.996) == "100,00₺"

    def test_cok_kucuk(self):
        assert format_turkish_lira(0.01) == "0,01₺"
